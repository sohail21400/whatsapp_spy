import selenium
from selenium import webdriver
from time import sleep
import openpyxl
from datetime import datetime
from datetime import timedelta

# NOTE: Timezone is not added
# least count is 10 seconds


def online_checker(user_name):

    options = webdriver.ChromeOptions()
    # location to the user profile data (create a new folder for that)
    options.add_argument("--user-data-dir=C:\\Users\\Sohail21400\\Desktop\\x")
    # add chrome web driver location here
    driver = webdriver.Chrome(
        executable_path='C:\\Users\\Sohail21400\\Downloads\\chromedriver_win32\\chromedriver.exe',
        options=options)

    work_book = openpyxl.load_workbook('online_status.xlsx')
    # NOTE: A SHEET SHOULD BE CREATED IN THE NAME OF user_name  BEFORE STARTING THIS PROGRAM
    sheet = work_book[user_name]
    try:
        driver.get("https://web.whatsapp.com/")
        is_main_page_loaded = False

        sleep(10)
        try_time = 3
        i = 0
        while not is_main_page_loaded and try_time > i:
            try:
                search_box = driver.find_element_by_xpath('//*[@id="side"]/div[1]/div/label/input')
                search_box.send_keys(user_name)
                # time to search user WE CAN ALSO IMPLEMENT A TRY EXCEPT BLOCK HERE BECAUSE THE SEARCH TIME MAY VARY
                sleep(2)
                user_icon = driver.find_element_by_xpath('//*[@id="pane-side"]/div[1]/div/div/div[2]')
                user_icon.click()
                # time to load user chat
                sleep(4)

                is_main_page_loaded = True

            except selenium.common.exceptions.NoSuchElementException:
                # this means the main WhatsApp page is not loaded yet
                print('Something went wrong.. trying again')
                # extended time to load the main page
                sleep(7)
            i += 1

        if not is_main_page_loaded:
            print("Couldn't load the website, check your connection and try again.")
            exit()

        # ------------------------------------------------------------------------------------------------------------------
        new_status = False
        old_status = True
        old_time = None
        time_period = None

        while True:  # Loop to continuously check the status
            try:  # if the user is offline the element won't be visible so trial and error
                is_online = driver.find_element_by_xpath('//*[@id="main"]/header/div[2]/div[2]/span').text
                if is_online == 'online':
                    new_status = True
                sleep(5)
            except selenium.common.exceptions.NoSuchElementException:
                new_status = False
                sleep(5)

            if old_status != new_status or old_time is None:
                new_time = datetime.now()
                if old_time is not None:
                    time_period = new_time - old_time
                old_time = new_time

                current_time = datetime.now().strftime("%H:%M:%S")
                current_date = datetime.now().strftime("%d:%m:%y")
                current_day = datetime.now().strftime("%a")

                # 1 --> Date
                # 2 --> Day
                # 3 --> Time
                # 4 --> Status
                # 5 --> Period
                max_row = sheet.max_row
                # argument order is (row, column)
                sheet.cell(max_row + 1, 1).value = current_date
                sheet.cell(max_row + 1, 2).value = current_day
                sheet.cell(max_row + 1, 3).value = current_time
                sheet.cell(max_row + 1, 4).value = new_status

                if time_period is not None:
                    sheet.cell(max_row + 1, 5).value = timedelta(seconds=time_period.seconds)
                else:
                    sheet.cell(max_row + 1, 5).value = 'None'

                old_status = new_status

            if new_status:
                print(f'{user_name} is online')
            else:
                print(f'{user_name} is offline')

    except selenium.common.exceptions.WebDriverException:

        work_book.save('online_status.xlsx')
        work_book.close()
        print('Webdriver is unreachable..')
        print('Ending process')


online_checker('Mom')





